from openpyxl import load_workbook

def readfile():

    wb =  load_workbook("list.xlsx")

    ws = wb['Elements']

    files_props = list()

    nun = 2
    while True:
        
        if ws[f'A{nun}'].value != None:
            
            a_temp = ws[f'A{nun}'].value
            b_temp = ws[f'B{nun}'].value if ws[f'B{nun}'].value != None else ""
            c_temp = ws[f'C{nun}'].value if ws[f'C{nun}'].value != None else "" 
            d_temp = ws[f'D{nun}'].value if ws[f'D{nun}'].value != None else ""  

            tup = (a_temp, b_temp, c_temp, d_temp)
            files_props.append(tup)   
            nun+=1
        else:
            break

    return files_props